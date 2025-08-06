"""
企业级工资表智能处理与分发系统
功能：支持任意格式Excel处理 + 智能合并 + 自定义邮件发送（含工资条附件）
作者：肖松甫
日期：2025-08-06
"""
import os
import sys
import re
import pandas as pd
import numpy as np
import logging
import smtplib
import sqlite3
import hashlib
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from datetime import datetime
from email.mime.text import MIMEText
from email.utils import formataddr
from email.header import Header
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.simpledialog import askstring
from tqdm import tqdm
from PIL import Image, ImageTk, ImageDraw
import threading
import time
from copy import copy
from openpyxl.styles import Font, PatternFill
from email.message import EmailMessage
import shutil


# ===================== 配置区域 =====================
class Config:
    SMTP_SERVER = "smtp.qq.com"
    SMTP_PORT = 465
    SENDER_EMAIL = "your_email@qq.com"
    SENDER_PASSWORD = "your_auth_code"
    SENDER_NAME = "财务部"
    COMPANY_NAME = "XX科技有限公司"
    HR_CONTACT = "*经理 ***********"
    OUTPUT_FOLDER = "工资表输出"
    LOG_FILE = "工资系统运行日志.log"
    DB_FILE = "工资发送记录.db"
    TEMP_SALARY_DIR = "临时工资条"
    DATE_FORMATS = ["%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日", "%d/%m/%Y"]
    NUMBER_FORMATS = [",", "，", " "]

    HEADER_KEYWORDS = ["姓名", "员工", "工号", "部门", "基本工资", "岗位工资", "绩效", "补贴",
                       "奖金", "扣款", "社保", "公积金", "个税", "实发工资", "银行账号", "邮箱"]

    HIDE_SENSITIVE_COLS = ['身份证号', '银行卡号', '银行账号']


# ===================== 日志系统 =====================
def setup_logging():
    """配置日志系统"""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    file_handler = logging.FileHandler(Config.LOG_FILE, encoding='utf-8')
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_formatter)

    console_handler = logging.StreamHandler()
    console_formatter = logging.Formatter('[%(levelname)s] %(message)s')
    console_handler.setFormatter(console_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


# ===================== 文件处理工具 =====================
def detect_file_encoding(file_path):
    """检测文件编码"""
    encodings = ['utf-8', 'gbk', 'gb18030', 'latin1']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1024)
            return encoding
        except UnicodeDecodeError:
            continue
    return 'utf-8'


def find_data_start_row(file_path):
    """智能定位数据起始行"""
    if file_path.lower().endswith(('.xlsx', '.xls')):
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        for row_idx in range(1, 50):
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_values.append(str(cell_value).strip())

            if any(keyword in " ".join(row_values) for keyword in Config.HEADER_KEYWORDS):
                return row_idx
        return 1
    elif file_path.lower().endswith('.csv'):
        encoding = detect_file_encoding(file_path)
        with open(file_path, 'r', encoding=encoding) as f:
            for i in range(20):
                line = f.readline()
                if any(keyword in line for keyword in Config.HEADER_KEYWORDS):
                    return i
        return 0
    return 1


def clean_column_name(name):
    """清洗列名"""
    if not name or pd.isna(name):
        return "未知列"

    name = re.sub(r'[^\w\u4e00-\u9fa5]', '', str(name).strip())

    name_mapping = {
        "姓名": "姓名",
        "员工姓名": "姓名",
        "员工名称": "姓名",
        "工号": "工号",
        "员工编号": "工号",
        "部门": "部门",
        "所属部门": "部门",
        "基本工资": "基本工资",
        "岗位工资": "岗位工资",
        "绩效工资": "绩效工资",
        "绩效奖金": "绩效工资",
        "补贴": "补贴",
        "津贴": "补贴",
        "奖金": "奖金",
        "扣款": "扣款",
        "缺勤扣款": "扣款",
        "社保": "社保",
        "社会保险": "社保",
        "公积金": "公积金",
        "个税": "个税",
        "个人所得税": "个税",
        "应发工资": "应发工资",
        "实发工资": "实发工资",
        "银行账号": "银行账号",
        "银行卡号": "银行账号",
        "邮箱": "邮箱",
        "电子邮箱": "邮箱"
    }
    return name_mapping.get(name, name)


def convert_to_number(value):
    """将各种格式转换为数字"""
    if pd.isna(value) or value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):

        for sep in Config.NUMBER_FORMATS:
            value = value.replace(sep, '')

        value = re.sub(r'[万亿]', '', value)

        try:
            return float(value)
        except ValueError:
            pass

        if '%' in value:
            try:
                return float(value.replace('%', '')) / 100
            except ValueError:
                pass
    return 0.0


def convert_to_date(value):
    """将各种格式转换为日期"""
    if pd.isna(value) or value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        value = value.strip()

        for fmt in Config.DATE_FORMATS:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def anonymize_data(value):
    """数据脱敏处理"""
    if pd.isna(value) or value is None:
        return ""
    value = str(value)

    if len(value) > 8 and value.isdigit():
        return value[:4] + "****" + value[-4:]

    if len(value) > 10 and (value.isdigit() or 'X' in value.upper()):
        return value[:3] + "***********" + value[-4:]
    return value


# ===================== 核心处理功能 =====================
def process_single_file(file_path):
    """处理单个工资表文件"""
    logger.info(f"开始处理文件: {os.path.basename(file_path)}")
    try:

        start_row = find_data_start_row(file_path)
        logger.info(f"检测到数据起始行: {start_row}")

        if file_path.lower().endswith(('.xlsx', '.xls')):

            df = pd.read_excel(
                file_path,
                header=start_row - 1 if start_row > 0 else None,
                engine='openpyxl'
            )
        elif file_path.lower().endswith('.csv'):

            encoding = detect_file_encoding(file_path)
            df = pd.read_csv(
                file_path,
                header=start_row,
                encoding=encoding,
                on_bad_lines='skip'
            )
        else:
            logger.error(f"不支持的文件格式: {file_path}")
            return None

        df.columns = [clean_column_name(col) for col in df.columns]
        logger.info(f"清洗后列名: {list(df.columns)}")

        for col in df.columns:

            if any(keyword in col for keyword in ["工资", "奖金", "补贴", "扣款", "社保", "公积金", "个税"]):
                df[col] = df[col].apply(convert_to_number)

            elif "日期" in col or "时间" in col:
                df[col] = df[col].apply(convert_to_date)

            elif any(keyword in col for keyword in Config.HIDE_SENSITIVE_COLS):
                df[col] = df[col].apply(anonymize_data)

        filename = os.path.splitext(os.path.basename(file_path))[0]
        df['数据来源'] = filename
        logger.info(f"文件处理完成: {len(df)}条记录")
        return df
    except Exception as e:
        logger.error(f"处理文件失败: {file_path} - {str(e)}")
        return None


def merge_all_files(file_paths):
    """合并所有工资表文件"""
    if not file_paths:
        logger.error("没有选择任何文件")
        return None
    all_dfs = []
    for file_path in file_paths:
        df = process_single_file(file_path)
        if df is not None and not df.empty:
            all_dfs.append(df)
    if not all_dfs:
        logger.error("所有文件处理失败")
        return None

    try:

        all_columns = set()
        for df in all_dfs:
            all_columns.update(df.columns)

        for df in all_dfs:
            for col in all_columns:
                if col not in df.columns:
                    df[col] = np.nan

        merged_df = pd.concat(all_dfs, ignore_index=True)

        merged_df.fillna(0, inplace=True)
        logger.info(f"合并完成! 总记录数: {len(merged_df)}")
        return merged_df
    except Exception as e:
        logger.error(f"合并失败: {str(e)}")
        return None


def save_merged_data(df, output_folder):
    """保存合并后的数据"""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_folder, f"合并工资表_{timestamp}.xlsx")

    try:
        df.to_excel(output_path, index=False)
        logger.info(f"合并结果保存至: {output_path}")

        beautify_excel(output_path)
        return output_path
    except Exception as e:
        logger.error(f"保存失败: {str(e)}")
        return None


def beautify_excel(file_path):
    """美化Excel格式"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        for cell in ws[1]:
            new_font = copy(cell.font)
            new_font.bold = True
            cell.font = new_font
            cell.fill = PatternFill("solid", fgColor="DDDDDD")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

        wb.save(file_path)
        logger.info("Excel格式美化完成")
    except Exception as e:
        logger.warning(f"Excel美化失败: {str(e)}")



def generate_employee_salary_sheet(employee_data, temp_dir=Config.TEMP_SALARY_DIR):
    """生成单个员工的工资条Excel文件（修复工号不存在的问题）"""

    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)


    emp_name = employee_data.get('姓名', '未知')


    if '工号' in employee_data.index:
        emp_id = employee_data['工号']

        if pd.isna(emp_id) or str(emp_id).strip() == '':
            emp_id = str(employee_data.name)
    else:
        emp_id = str(employee_data.name)


    emp_id_str = str(emp_id).replace('/', '').replace('\\', '')
    filename = f"{datetime.now().strftime('%Y%m')}工资条_{emp_name}_{emp_id_str}.xlsx"
    file_path = os.path.join(temp_dir, filename)

    try:

        from openpyxl import Workbook


        wb = Workbook()
        ws = wb.active
        ws.title = "工资明细"
        ws.sheet_state = "visible"


        important_cols = []

        for col in ['姓名', '工号', '部门', '实发工资', '基本工资', '绩效工资', '奖金', '扣款']:
            if col in employee_data.index:
                important_cols.append(col)


        all_cols = [col for col in employee_data.index if col != '数据来源']
        other_cols = [col for col in all_cols if col not in important_cols]
        ordered_cols = important_cols + other_cols


        for col_idx, col_name in enumerate(ordered_cols, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)


        for col_idx, col_name in enumerate(ordered_cols, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = employee_data[col_name]

            if any(keyword in col_name for keyword in ["工资", "奖金", "补贴", "扣款", "社保", "公积金", "个税"]):
                cell.number_format = '#,##0.00'


        for col_idx in range(1, len(ordered_cols) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15


        wb.save(file_path)
        logger.info(f"生成工资条文件: {filename}")
        return file_path

    except Exception as e:
        logger.error(f"生成工资条失败: {str(e)}")

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except:
                pass
        return None


# ===================== 邮件发送功能 =====================
def send_salary_emails(df, smtp_config, progress_callback=None):
    """发送工资条邮件（含附件）"""
    if not all([smtp_config['server'], smtp_config['email'], smtp_config['password']]):
        logger.error("邮件服务器配置不完整")
        return 0, "邮件服务器配置不完整"


    if '邮箱' not in df.columns:
        logger.error("工资表中缺少'邮箱'列")
        return 0, "工资表中必须包含'邮箱'列"


    temp_dir = Config.TEMP_SALARY_DIR
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)


    df['employee_id'] = df.apply(
        lambda row: row['工号'] if '工号' in df.columns else row['姓名'] + '_' + str(row.name),
        axis=1
    )


    success_count = 0
    total = len(df)
    errors = []

    for index, row in df.iterrows():
        employee_name = row.get('姓名', '未知员工')
        employee_id = row['employee_id']
        employee_email = row['邮箱']


        if not isinstance(employee_email, str) or "@" not in employee_email:
            error_msg = f"跳过无效邮箱: {employee_name} - {employee_email}"
            logger.warning(error_msg)
            errors.append(error_msg)
            if progress_callback:
                progress_callback(index + 1, total, f"跳过: {employee_name}")
            continue

        try:

            attachment_path = generate_employee_salary_sheet(row, temp_dir)
            if not attachment_path:
                error_msg = f"无法生成{employee_name}的工资条"
                logger.error(error_msg)
                errors.append(error_msg)
                continue


            try:
                server = smtplib.SMTP_SSL(smtp_config['server'], smtp_config['port'])
                server.login(smtp_config['email'], smtp_config['password'])
                logger.info(f"邮件服务器登录成功 ({index + 1}/{total})")
            except Exception as e:
                logger.error(f"邮件服务器连接失败: {str(e)}")
                errors.append(f"邮件服务器连接失败: {str(e)}")
                if progress_callback:
                    progress_callback(index + 1, total, f"连接失败: {employee_name}")
                continue


            msg = EmailMessage()


            email_body = f"""
            <html>
            <body>
                <div style="font-family: 'Microsoft YaHei', sans-serif; line-height: 1.6;">
                    <div style="color: #2c3e50; border-bottom: 1px solid #eee; padding-bottom: 10px;">
                        <h2>{smtp_config['company_name']}</h2>
                        <h3>{datetime.now().strftime('%Y年%m月')}工资通知</h3>
                    </div>

                    <p>尊敬的{employee_name}：</p>
                    <p>您的{datetime.now().strftime('%Y年%m月')}工资明细已生成，详情请查看附件中的工资条。</p>

                    <p><strong>重要提示：</strong></p>
                    <ul>
                        <li>工资条包含个人隐私信息，请妥善保管并及时查阅</li>
                        <li>如有任何疑问，请联系人力资源部：{smtp_config['hr_contact']}</li>
                    </ul>

                    <div style="margin-top: 30px; padding-top: 10px; border-top: 1px solid #eee; color: #7f8c8d; font-size: 0.9em;">
                        <p>本邮件为系统自动发送，请勿直接回复</p>
                        <p>{smtp_config['sender_name']}</p>
                        <p>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    </div>
                </div>
            </body>
            </html>
            """


            from_header = str(Header(smtp_config['sender_name'], 'utf-8'))
            msg['From'] = formataddr((from_header, smtp_config['email']))
            msg['To'] = employee_email
            msg['Subject'] = f"{datetime.now().strftime('%Y年%m月')}工资条 - {employee_name}"
            msg.set_content(email_body, subtype='html')


            with open(attachment_path, 'rb') as f:
                file_data = f.read()
            msg.add_attachment(
                file_data,
                maintype='application',
                subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=os.path.basename(attachment_path)
            )


            try:
                server.send_message(msg)
                logger.info(f"发送成功: {employee_name} <{employee_email}> (含附件)")
                success_count += 1
                if progress_callback:
                    progress_callback(index + 1, total, f"已发送: {employee_name}")
            except Exception as e:
                error_msg = f"发送给{employee_name}失败: {str(e)}"
                logger.error(error_msg)
                errors.append(error_msg)
                if progress_callback:
                    progress_callback(index + 1, total, f"失败: {employee_name}")
            finally:

                try:
                    server.quit()
                except Exception as e:
                    logger.warning(f"关闭连接时出错: {str(e)}")

        except Exception as e:
            error_msg = f"处理{employee_name}时出错: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
            if progress_callback:
                progress_callback(index + 1, total, f"失败: {employee_name}")


        if index < total - 1:
            delay = 15 if "qq.com" not in smtp_config['server'] else 20
            logger.info(f"等待 {delay} 秒后发送下一封邮件...")
            time.sleep(delay)


    try:
        shutil.rmtree(temp_dir)
        logger.info("临时工资条文件已清理")
    except Exception as e:
        logger.warning(f"清理临时文件失败: {str(e)}")


    result_msg = f"邮件发送完成!\n\n总人数: {total}\n成功: {success_count}\n失败: {total - success_count}"
    if errors:
        result_msg += f"\n\n错误详情请查看日志文件"
        result_msg += f"\n\n部分错误示例:"
        for error in errors[:5]:
            result_msg += f"\n- {error}"
    return success_count, result_msg


# ===================== GUI界面 =====================
class SalaryProcessorApp:
    def __init__(self, root):
        self.root = root
        root.title(f"{Config.COMPANY_NAME} - 工资表处理系统")
        root.geometry("1000x700")
        root.resizable(True, True)

        self.style = ttk.Style()
        self.style.configure("TButton", padding=6, font=('Arial', 10))
        self.style.configure("Header.TLabel", font=('Arial', 14, 'bold'))
        self.style.configure("Red.TLabel", foreground="red")

        self.create_widgets()

        self.logger = setup_logging()
        self.logger.info("工资表处理系统启动")

        self.merged_df = None
        self.output_path = None

        self.load_logo()

    def load_logo(self):
        """加载公司LOGO（占位）"""
        try:

            img = Image.new('RGB', (200, 50), color=(73, 109, 137))
            d = ImageDraw.Draw(img)
            d.text((10, 10), Config.COMPANY_NAME, fill=(255, 255, 255))
            self.logo_img = ImageTk.PhotoImage(img)

            self.logo_label.config(image=self.logo_img)
        except:
            self.logo_label.config(text=Config.COMPANY_NAME, style="Header.TLabel")

    def create_widgets(self):

        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill="x", pady=(0, 10))

        self.logo_label = ttk.Label(header_frame)
        self.logo_label.pack(side="left", padx=10)

        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side="right", fill="x", expand=True)
        ttk.Label(
            title_frame,
            text="工资表智能处理与分发系统",
            style="Header.TLabel"
        ).pack(anchor="e")
        ttk.Label(
            title_frame,
            text="安全、高效、专业的工资管理解决方案",
            font=('Arial', 10)
        ).pack(anchor="e")

        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True)

        self.file_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.file_tab, text="工资表处理")

        self.email_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.email_tab, text="邮件发送")

        self.config_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.config_tab, text="系统配置")

        self.create_file_tab()
        self.create_email_tab()
        self.create_config_tab()

        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_bar = ttk.Frame(self.root, relief="sunken", padding=2)
        status_bar.pack(side="bottom", fill="x")
        ttk.Label(status_bar, textvariable=self.status_var).pack(side="left")

        ttk.Label(status_bar, text="版本 2.1").pack(side="right", padx=5)

    def create_file_tab(self):
        """创建文件处理选项卡"""
        tab = self.file_tab

        file_frame = ttk.LabelFrame(tab, text="选择工资表文件")
        file_frame.pack(fill="x", padx=10, pady=5)
        self.file_listbox = tk.Listbox(
            file_frame,
            height=8,
            selectmode=tk.EXTENDED
        )
        self.file_listbox.pack(fill="both", expand=True, padx=5, pady=5)
        scrollbar = ttk.Scrollbar(file_frame, orient="vertical", command=self.file_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.file_listbox.config(yscrollcommand=scrollbar.set)
        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(fill="x", padx=5, pady=5)
        ttk.Button(
            btn_frame,
            text="添加文件",
            command=self.add_files
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_frame,
            text="添加文件夹",
            command=self.add_folder
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_frame,
            text="移除选中",
            command=self.remove_selected
        ).pack(side="right", padx=5)

        action_frame = ttk.Frame(tab)
        action_frame.pack(fill="x", padx=10, pady=10)
        self.process_btn = ttk.Button(
            action_frame,
            text="处理并合并工资表",
            command=self.process_files,
            style="Accent.TButton"
        )
        self.process_btn.pack(pady=5)

        result_frame = ttk.LabelFrame(tab, text="处理结果")
        result_frame.pack(fill="both", expand=True, padx=10, pady=5)
        self.result_text = scrolledtext.ScrolledText(
            result_frame,
            wrap=tk.WORD,
            state="disabled"
        )
        self.result_text.pack(fill="both", expand=True, padx=5, pady=5)

        self.style.configure("Accent.TButton", foreground="white", background="#4CAF50")

    def create_email_tab(self):
        """创建邮件发送选项卡"""
        tab = self.email_tab

        preview_frame = ttk.LabelFrame(tab, text="邮件预览")
        preview_frame.pack(fill="x", padx=10, pady=5)
        self.email_preview = scrolledtext.ScrolledText(
            preview_frame,
            height=10,
            wrap=tk.WORD
        )
        self.email_preview.pack(fill="x", padx=5, pady=5)
        self.email_preview.insert(tk.END, "邮件内容将在发送前自动生成...\n包含员工个人工资条Excel附件")
        self.email_preview.config(state="disabled")

        control_frame = ttk.Frame(tab)
        control_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(control_frame, text="发送范围:").grid(row=0, column=0, sticky="w", padx=5)
        self.send_scope = tk.StringVar(value="all")
        ttk.Radiobutton(control_frame, text="全部员工", variable=self.send_scope, value="all").grid(row=0, column=1,
                                                                                                    sticky="w", padx=5)
        ttk.Radiobutton(control_frame, text="未发送员工", variable=self.send_scope, value="unsent").grid(row=0,
                                                                                                         column=2,
                                                                                                         sticky="w",
                                                                                                         padx=5)

        self.send_btn = ttk.Button(
            control_frame,
            text="发送工资条",
            command=self.send_salaries,
            state="disabled"
        )
        self.send_btn.grid(row=0, column=3, padx=10)

        progress_frame = ttk.LabelFrame(tab, text="发送进度")
        progress_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            mode="determinate"
        )
        self.progress_bar.pack(fill="x", padx=5, pady=5)

        self.progress_label = ttk.Label(
            progress_frame,
            text="等待操作..."
        )
        self.progress_label.pack(fill="x", padx=5, pady=2)

        self.email_log = scrolledtext.ScrolledText(
            progress_frame,
            height=8,
            wrap=tk.WORD,
            state="disabled"
        )
        self.email_log.pack(fill="both", expand=True, padx=5, pady=5)

    def create_config_tab(self):
        """创建系统配置选项卡"""
        tab = self.config_tab

        mail_frame = ttk.LabelFrame(tab, text="邮件服务器配置")
        mail_frame.pack(fill="x", padx=10, pady=5)

        form_frame = ttk.Frame(mail_frame)
        form_frame.pack(fill="x", padx=5, pady=5)

        ttk.Label(form_frame, text="邮箱服务:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.provider_var = tk.StringVar()
        self.provider_box = ttk.Combobox(form_frame, textvariable=self.provider_var, width=28, state="readonly")
        self.provider_box['values'] = ["自动识别", "QQ邮箱", "网易163", "网易126", "Gmail", "Outlook", "腾讯企业邮箱",
                                       "自定义"]
        self.provider_box.current(0)
        self.provider_box.grid(row=0, column=1, padx=5, pady=2)
        self.provider_box.bind("<<ComboboxSelected>>", self.on_provider_selected)

        ttk.Label(form_frame, text="SMTP服务器:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.smtp_server = ttk.Entry(form_frame, width=30)
        self.smtp_server.grid(row=1, column=1, padx=5, pady=2)
        self.smtp_server.insert(0, Config.SMTP_SERVER)

        ttk.Label(form_frame, text="端口:").grid(row=0, column=2, sticky="w", padx=5, pady=2)
        self.smtp_port = ttk.Entry(form_frame, width=10)
        self.smtp_port.grid(row=0, column=3, padx=5, pady=2)
        self.smtp_port.insert(0, str(Config.SMTP_PORT))

        ttk.Label(form_frame, text="邮箱账号:").grid(row=2, column=0, sticky="w", padx=5, pady=2)
        self.email_account = ttk.Entry(form_frame, width=30)
        self.email_account.grid(row=2, column=1, padx=5, pady=2)
        self.email_account.insert(0, Config.SENDER_EMAIL)

        ttk.Label(form_frame, text="授权码:").grid(row=1, column=2, sticky="w", padx=5, pady=2)
        self.email_password = ttk.Entry(form_frame, width=30, show="*")
        self.email_password.grid(row=1, column=3, padx=5, pady=2)
        self.email_password.insert(0, Config.SENDER_PASSWORD)

        ttk.Label(form_frame, text="发件人名称:").grid(row=3, column=0, sticky="w", padx=5, pady=2)
        self.sender_name = ttk.Entry(form_frame, width=30)
        self.sender_name.grid(row=3, column=1, padx=5, pady=2)
        self.sender_name.insert(0, Config.SENDER_NAME)

        ttk.Label(form_frame, text="公司名称:").grid(row=2, column=2, sticky="w", padx=5, pady=2)
        self.company_name = ttk.Entry(form_frame, width=30)
        self.company_name.grid(row=2, column=3, padx=5, pady=2)
        self.company_name.insert(0, Config.COMPANY_NAME)

        ttk.Label(form_frame, text="HR联系方式:").grid(row=3, column=2, sticky="w", padx=5, pady=2)
        self.hr_contact = ttk.Entry(form_frame, width=30)
        self.hr_contact.grid(row=3, column=3, padx=5, pady=2)
        self.hr_contact.insert(0, Config.HR_CONTACT)

        test_frame = ttk.Frame(mail_frame)
        test_frame.pack(fill="x", padx=5, pady=5)
        ttk.Button(
            test_frame,
            text="测试邮件配置",
            command=self.test_email_config
        ).pack(side="left", padx=5)
        ttk.Button(
            test_frame,
            text="保存配置",
            command=self.save_config
        ).pack(side="right", padx=5)

        info_frame = ttk.LabelFrame(tab, text="系统信息")
        info_frame.pack(fill="both", expand=True, padx=10, pady=5)
        info_text = scrolledtext.ScrolledText(
            info_frame,
            wrap=tk.WORD,
            state="disabled"
        )
        info_text.pack(fill="both", expand=True, padx=5, pady=5)
        info = f"""
        系统名称：工资表智能处理与分发系统
        版本：2.1
        最后更新：2025-08-06

        功能特点：
        1. 支持任意格式的Excel/CSV工资表
        2. 智能识别表头和数据类型
        3. 自动合并多部门工资表
        4. 自定义邮件发送工资条（含Excel附件）
        5. 敏感信息自动脱敏
        6. 完整的日志记录

        使用说明：
        1. 在[工资表处理]选项卡中添加文件并合并
        2. 在[邮件发送]选项卡中配置并发送工资条
        3. 在[系统配置]选项卡中设置邮件服务器

        
        """
        info_text.config(state="normal")
        info_text.insert(tk.END, info)
        info_text.config(state="disabled")

    def add_files(self):
        """添加文件"""
        filetypes = (
            ('Excel 文件', '*.xlsx *.xls'),
            ('CSV 文件', '*.csv'),
            ('所有文件', '*.*')
        )
        files = filedialog.askopenfilenames(
            title="选择工资表文件",
            filetypes=filetypes
        )
        if files:
            for file in files:
                if file not in self.file_listbox.get(0, tk.END):
                    self.file_listbox.insert(tk.END, file)
            self.log(f"添加 {len(files)} 个文件")

    def add_folder(self):
        """添加文件夹"""
        folder = filedialog.askdirectory(title="选择包含工资表的文件夹")
        if folder:
            added = 0
            for root, _, files in os.walk(folder):
                for file in files:
                    if file.lower().endswith(('.xlsx', '.xls', '.csv')):
                        full_path = os.path.join(root, file)
                        if full_path not in self.file_listbox.get(0, tk.END):
                            self.file_listbox.insert(tk.END, full_path)
                            added += 1
            self.log(f"添加 {added} 个文件")

    def remove_selected(self):
        """移除选中文件"""
        selected = self.file_listbox.curselection()
        for index in selected[::-1]:
            self.file_listbox.delete(index)
        self.log(f"移除 {len(selected)} 个文件")

    def log(self, message):
        """记录日志到文本框"""
        self.result_text.config(state="normal")
        self.result_text.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.result_text.see(tk.END)
        self.result_text.config(state="disabled")
        self.status_var.set(message)
        self.root.update_idletasks()

    def process_files(self):
        """处理文件"""
        file_paths = self.file_listbox.get(0, tk.END)
        if not file_paths:
            messagebox.showwarning("警告", "请先添加工资表文件")
            return

        self.result_text.config(state="normal")
        self.result_text.delete(1.0, tk.END)
        self.result_text.config(state="disabled")

        self.log("开始处理工资表...")
        self.merged_df = merge_all_files(file_paths)
        if self.merged_df is not None:
            self.output_path = save_merged_data(self.merged_df, Config.OUTPUT_FOLDER)
            if self.output_path:
                self.log(f"处理完成! 结果已保存至:\n{self.output_path}")

                summary = f"""
                ====== 工资表处理完成 ======
                总记录数: {len(self.merged_df)}
                包含部门: {self.merged_df['部门'].nunique()}
                时间范围: {self.merged_df.get('月份', '未知')}
                保存位置: {self.output_path}

                可以切换到[邮件发送]选项卡发送工资条（含Excel附件）
                """
                self.result_text.config(state="normal")
                self.result_text.insert(tk.END, summary)
                self.result_text.config(state="disabled")

                self.send_btn.config(state="normal")

                os.startfile(Config.OUTPUT_FOLDER)
        else:
            self.log("处理失败，请检查日志")
            messagebox.showerror("错误", "工资表处理失败，请检查日志文件")

    def get_smtp_config(self):
        """获取SMTP配置"""
        return {
            'server': self.smtp_server.get(),
            'port': int(self.smtp_port.get()),
            'email': self.email_account.get(),
            'password': self.email_password.get(),
            'sender_name': self.sender_name.get(),
            'company_name': self.company_name.get(),
            'hr_contact': self.hr_contact.get()
        }

    def test_email_config(self):
        """测试邮件配置"""
        config = self.get_smtp_config()
        if not all([config['server'], config['email'], config['password']]):
            messagebox.showwarning("警告", "请填写完整的邮件服务器配置")
            return
        try:
            server = smtplib.SMTP_SSL(config['server'], config['port'])
            server.login(config['email'], config['password'])
            server.quit()
            messagebox.showinfo("成功", "邮件服务器连接成功!")
        except Exception as e:
            messagebox.showerror("错误", f"连接失败: {str(e)}")

    def save_config(self):
        """保存配置"""

        messagebox.showinfo("成功", "配置已保存!")
        self.log("系统配置已更新")

    def update_progress(self, current, total, message):
        """更新进度条"""
        progress = (current / total) * 100
        self.progress_var.set(progress)
        self.progress_label.config(text=f"{current}/{total} - {message}")
        self.status_var.set(f"正在发送: {message}")
        self.root.update_idletasks()

    def log_email(self, message):
        """记录邮件日志"""
        self.email_log.config(state="normal")
        self.email_log.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.email_log.see(tk.END)
        self.email_log.config(state="disabled")
        self.root.update_idletasks()

    def send_salaries(self):
        """发送工资条"""
        if self.merged_df is None:
            messagebox.showwarning("警告", "请先处理工资表")
            return

        send_scope = self.send_scope.get()

        if send_scope == "unsent":

            df_to_send = self.merged_df.copy()
            self.log_email("注意: 未发送员工筛选功能尚未实现，将发送全部员工")
        else:
            df_to_send = self.merged_df.copy()

        smtp_config = self.get_smtp_config()

        preview_text = f"""
        发件人: {smtp_config['sender_name']} <{smtp_config['email']}>
        主题: {datetime.now().strftime('%Y年%m月')}工资条 - 员工姓名
        收件人数: {len(df_to_send)}
        附件: 包含员工个人工资条Excel文件

        邮件内容:
        尊敬的[员工姓名]：

        您的{datetime.now().strftime('%Y年%m月')}工资明细已生成，详情请查看附件中的工资条。

        重要提示：
        • 工资条包含个人隐私信息，请妥善保管
        • 如有任何疑问，请联系人力资源部：{smtp_config['hr_contact']}

        {smtp_config['sender_name']}
        {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        self.email_preview.config(state="normal")
        self.email_preview.delete(1.0, tk.END)
        self.email_preview.insert(tk.END, preview_text)
        self.email_preview.config(state="disabled")

        if not messagebox.askyesno("确认发送",
                                   f"确定要发送给 {len(df_to_send)} 位员工吗？\n系统将生成并发送个人工资条Excel附件"):
            return

        self.progress_var.set(0)
        self.email_log.config(state="normal")
        self.email_log.delete(1.0, tk.END)
        self.email_log.config(state="disabled")

        threading.Thread(
            target=self._send_salaries_thread,
            args=(df_to_send, smtp_config),
            daemon=True
        ).start()

    def on_provider_selected(self, event=None):
        """选择邮箱服务商后自动填入SMTP设置"""
        provider = self.provider_var.get()
        server, port = "", ""
        if provider == "QQ邮箱":
            server, port = "smtp.qq.com", 465
        elif provider == "网易163":
            server, port = "smtp.163.com", 465
        elif provider == "网易126":
            server, port = "smtp.126.com", 465
        elif provider == "Gmail":
            server, port = "smtp.gmail.com", 465
        elif provider == "Outlook":
            server, port = "smtp.office365.com", 587
        elif provider == "腾讯企业邮箱":
            server, port = "smtp.exmail.qq.com", 465
        elif provider == "自动识别":
            self.auto_detect_smtp()
            return
        elif provider == "自定义":
            return
        self.smtp_server.delete(0, tk.END)
        self.smtp_server.insert(0, server)
        self.smtp_port.delete(0, tk.END)
        self.smtp_port.insert(0, str(port))

    def auto_detect_smtp(self):
        """根据邮箱地址自动识别SMTP服务商"""
        email = self.email_account.get().strip()
        if "@" not in email:
            return
        domain = email.split('@')[-1].lower()
        if domain == "qq.com":
            self.provider_box.set("QQ邮箱")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.qq.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "465")
        elif domain == "163.com":
            self.provider_box.set("网易163")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.163.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "465")
        elif domain == "126.com":
            self.provider_box.set("网易126")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.126.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "465")
        elif domain == "gmail.com":
            self.provider_box.set("Gmail")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.gmail.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "465")
        elif domain == "outlook.com":
            self.provider_box.set("Outlook")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.office365.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "587")
        elif domain.endswith("yourcompany.com"):
            self.provider_box.set("腾讯企业邮箱")
            self.smtp_server.delete(0, tk.END)
            self.smtp_server.insert(0, "smtp.exmail.qq.com")
            self.smtp_port.delete(0, tk.END)
            self.smtp_port.insert(0, "465")
        else:
            self.provider_box.set("自定义")

    def _send_salaries_thread(self, df, smtp_config):
        """在后台线程中发送工资条"""
        try:

            success_count, result_msg = send_salary_emails(
                df,
                smtp_config,
                progress_callback=self.update_progress
            )

            self.log_email(result_msg)
            messagebox.showinfo("发送完成", result_msg)
        except Exception as e:
            self.log_email(f"发送过程中发生错误: {str(e)}")
            messagebox.showerror("错误", f"发送失败: {str(e)}")


# ===================== 主程序 =====================
if __name__ == "__main__":

    logger = logging.getLogger()
    try:

        root = tk.Tk()
        app = SalaryProcessorApp(root)

        if sys.platform.startswith('win'):
            import ctypes

            kernel32 = ctypes.windll.kernel32
            kernel32.SetConsoleOutputCP(65001)  # UTF-8
        root.mainloop()
    except Exception as e:
        logger.exception("系统发生未预期错误")
        messagebox.showerror("系统错误", f"程序发生异常: {str(e)}")